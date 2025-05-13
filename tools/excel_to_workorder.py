import os
import pymysql
import openpyxl
import sys

# 数据库连接配置
MYSQL_CONFIG = {
    'host': '221.228.110.148',
    'user': 'root',
    'password': 'thriftdy@819',
    'database': 'test',
    'port': 3306,
    'charset': 'utf8mb4'
}

# Excel列名 : workorder表字段名
# 请根据实际Excel表头和数据库字段自定义映射关系
COLUMN_FIELD_MAP = {
     '业务分类': 'busitype',
     '#': 'orderno',
     '跟踪': 'follow',
     '状态': 'state',
     '主题': 'subject',
     '作者': 'author',
     '风险星级': 'risk',
     '预期时间': 'workhours',
     '缺陷数量': 'bugs',
     '计划提测日期': 'submittest',
     '计划完成日期': 'plandate',
     '实际完成日期': 'actualdate',
     '产品人员': 'productor',
     '技术人员': 'tech',
     '测试人员': 'tester',
     '质控分值': 'level',
     '速度分值': 'speed',
     '质量分值': 'quality',
     '态度分值': 'attitude',
     # '月份': 'tmonth',  # 由参数传入
}


def insert_to_workorder(data, fields):
    if not data:
        print('没有可插入的数据。')
        return
    conn = pymysql.connect(**MYSQL_CONFIG)
    cursor = conn.cursor()
    placeholders = ','.join(['%s'] * len(fields))
    sql = f"INSERT INTO workorder ({','.join(fields)}) VALUES ({placeholders})"
    cursor.executemany(sql, data)
    conn.commit()
    cursor.close()
    conn.close()


def read_excel_and_insert(file_path, tmonth=None):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(cell).strip() for cell in rows[0]]
    print(f"Excel表头如下：{header}")
    if not COLUMN_FIELD_MAP:
        print("请在 COLUMN_FIELD_MAP 中配置映射关系后再运行！")
        return
    # 只取映射中存在的列
    field_indices = {COLUMN_FIELD_MAP[col]: header.index(col) for col in COLUMN_FIELD_MAP if col in header}
    fields = list(field_indices.keys())
    if tmonth is not None:
        fields.append('tmonth')
    data = []
    for row in rows[1:]:
        record = []
        for field in field_indices:
            idx = field_indices[field]
            record.append(row[idx] if idx is not None else None)
        if tmonth is not None:
            record.append(tmonth)
        data.append(tuple(record))
    insert_to_workorder(data, fields)
    print(f"已插入 {len(data)} 条数据到 workorder 表。")


def batch_import_from_dir(dir_path, tmonth=None):
    for fname in os.listdir(dir_path):
        if fname.endswith('.xlsx'):
            file_path = os.path.join(dir_path, fname)
            print(f"正在处理文件: {file_path}")
            read_excel_and_insert(file_path, tmonth)


def get_workorder_fields():
    conn = pymysql.connect(**MYSQL_CONFIG)
    cursor = conn.cursor()
    cursor.execute("DESC workorder;")
    fields = [row[0] for row in cursor.fetchall() if row[0] != 'id']
    cursor.close()
    conn.close()
    return fields


def get_excel_header(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header = [str(cell).strip() for cell in next(ws.iter_rows(values_only=True))]
    return header

def get_excel_header_and_workorder_fields():
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx')]
    if not excel_files:
        print("未找到Excel文件")
        exit(1)
    excel_path = os.path.join(excel_dir, excel_files[0])
    excel_header = get_excel_header(excel_path)
    workorder_fields = get_workorder_fields()

    print("Excel表头：", excel_header)
    print("workorder表字段：", workorder_fields)

    # 自动生成建议映射关系
    print("\n建议的 COLUMN_FIELD_MAP 映射关系模板：\nCOLUMN_FIELD_MAP = {")
    for col in excel_header:
        # 如果表头和字段名一致或相似，自动匹配
        match = next((f for f in workorder_fields if f.lower() == col.lower()), '')
        if match:
            print(f"    '{col}': '{match}',")
        else:
            print(f"    # '{col}': '',")
    print("}") 

if __name__ == '__main__':
    excel_dir = os.path.join('source', 'workorder')
    tmonth = '202504'
    if len(sys.argv) > 1:
        tmonth = sys.argv[1]
    batch_import_from_dir(excel_dir, tmonth)
