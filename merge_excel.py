import openpyxl
from pathlib import Path
import os
from copy import copy
import win32com.client
import time
import xlwings as xw

def convert_xls_to_xlsx(xls_path):
    """使用 xlwings 将 xls 转换为 xlsx，保留所有格式"""
    app = None
    wb = None
    
    try:
        # 获取完整路径
        abs_path = str(xls_path.absolute())
        xlsx_path = str(xls_path.parent / f"temp_{xls_path.stem}.xlsx")
        
        print(f"开始转换文件: {xls_path.name}")
        
        # 如果目标文件已存在，先删除
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        
        # 启动Excel应用程序
        app = xw.App(visible=False)
        app.display_alerts = False
        
        # 打开工作簿
        wb = app.books.open(abs_path)
        
        # 另存为xlsx格式
        wb.save(xlsx_path)
        
        # 关闭工作簿
        wb.close()
        
        return Path(xlsx_path)
        
    except Exception as e:
        print(f"转换文件时出错 {xls_path}: {str(e)}")
        return None
        
    finally:
        # 确保资源被正确释放
        try:
            if wb is not None:
                wb.close()
        except:
            pass
            
        try:
            if app is not None:
                app.quit()
        except:
            pass
        
def merge_summary_sheets(summary_sheets):
    """合并汇总表，正确处理表头"""
    if not summary_sheets:
        return None
        
    try:
        # 获取第一个sheet作为基准
        first_sheet = summary_sheets[0]
        header_row = None
        merged_rows = []
        
        # 从第一个sheet获取表头
        for row in first_sheet.rows:
            header_row = [cell.value for cell in row]
            break  # 只获取第一行作为表头
            
        if not header_row:
            return None
            
        # 处理每个汇总sheet
        for sheet in summary_sheets:
            is_first_row = True
            for row in sheet.rows:
                # 跳过每个sheet的表头行（第一行）
                if is_first_row:
                    is_first_row = False
                    continue
                    
                # 收集行数据
                row_data = []
                for cell in row:
                    row_data.append({
                        'value': cell.value,
                        'cell': cell  # 保存原始单元格以便复制格式
                    })
                merged_rows.append(row_data)
        
        return {
            'header': header_row,
            'rows': merged_rows,
            'header_format': list(first_sheet.rows)[0]  # 保存表头行的格式
        }
        
    except Exception as e:
        print(f"合并汇总表时出错: {str(e)}")
        return None

def write_merged_summary(workbook, merged_data):
    """将合并后的汇总数据写入新的工作簿"""
    if not merged_data or 'header' not in merged_data:
        return
        
    try:
        # 创建汇总sheet
        summary_sheet = workbook.create_sheet(title='总表', index=0)
        
        # 写入表头并复制格式
        for col_idx, header_value in enumerate(merged_data['header'], 1):
            cell = summary_sheet.cell(row=1, column=col_idx, value=header_value)
            # 复制表头格式
            source_cell = merged_data['header_format'][col_idx-1]
            copy_cell_format(source_cell, cell)
        
        # 写入数据行
        for row_idx, row_data in enumerate(merged_data['rows'], 2):  # 从第2行开始
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = summary_sheet.cell(
                    row=row_idx,
                    column=col_idx,
                    value=cell_data['value']
                )
                # 复制单元格格式
                copy_cell_format(cell_data['cell'], cell)
                
    except Exception as e:
        print(f"写入合并后的汇总表时出错: {str(e)}")
        
def copy_cell_format(source_cell, target_cell):
    """复制单元格格式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def is_valid_excel(file_path):
    """验证Excel文件是否有效"""
    try:
        if file_path.suffix.lower() == '.xls':
            # 尝试用 Excel COM 对象打开文件
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(str(file_path.absolute()))
            wb.Close()
            excel.Quit()
        else:
            openpyxl.load_workbook(file_path, read_only=True)
        return True
    except Exception as e:
        print(f"无效的Excel文件 {file_path}: {str(e)}")
        return False

def merge_excel_files():
    # 定义源目录路径，用于存放待合并的Excel文件
    source_dir = Path('source/performance')
    
    # 定义输出目录路径，用于存放合并后的Excel文件
    output_dir = Path('output')
    
    # 如果输出目录不存在，则创建该目录
    output_dir.mkdir(exist_ok=True)
    
    # 初始化一个列表，用于存储临时文件路径
    temp_files = []
    
    # 创建新的工作簿
    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)
    
    # 用于存储所有"汇总"sheet的数据和格式
    summary_cells = []
    
    try:
        # 获取所有Excel文件
        excel_files = list(source_dir.glob('*.xlsx')) + list(source_dir.glob('*.xls'))
        if not excel_files:
            raise Exception("source目录中没有找到Excel文件")
        
        # 用于存储所有汇总sheet
        summary_sheets = []

        # 遍历source目录下的所有excel文件
        for excel_file in excel_files:
            print(f"正在处理文件: {excel_file.name}")
            
            # 验证文件是否有效
            if not is_valid_excel(excel_file):
                print(f"跳过无效文件: {excel_file.name}")
                continue
            
            # 如果是xls文件，先转换为xlsx
            if excel_file.suffix.lower() == '.xls':
                print(f"转换 {excel_file.name} 为xlsx格式")
                temp_xlsx = convert_xls_to_xlsx(excel_file)
                if temp_xlsx is None:
                    continue
                excel_file = temp_xlsx
                temp_files.append(temp_xlsx)
                
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
            except Exception as e:
                print(f"无法打开文件 {excel_file.name}: {str(e)}")
                continue
            
            # 处理"汇总"sheet
            if '总表' in wb.sheetnames :
                print(f"正在处理 {excel_file.name} 的总表sheet")
                summary_sheets.append(wb['总表'])
               
            
            # 复制其他sheet
            for sheet_name in wb.sheetnames:
                if sheet_name != '总表':
                    print(f"正在复制 {excel_file.name} 的 {sheet_name} sheet")
                    source_sheet = wb[sheet_name]
                    # 创建新的sheet，添加序号避免重名
                    #new_sheet_name = f"{sheet_name}_{len(merged_wb.sheetnames)}"
                    target_sheet = merged_wb.create_sheet(title=sheet_name)
                    
                    # 复制列宽
                    for col in range(1, source_sheet.max_column + 1):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        if col_letter in source_sheet.column_dimensions:
                            target_sheet.column_dimensions[col_letter].width = \
                                source_sheet.column_dimensions[col_letter].width
                    
                    # 复制行高
                    for row in range(1, source_sheet.max_row + 1):
                        if row in source_sheet.row_dimensions:
                            target_sheet.row_dimensions[row].height = \
                                source_sheet.row_dimensions[row].height
                    
                    # 复制单元格内容和格式
                    for row in source_sheet.rows:
                        for cell in row:
                            new_cell = target_sheet.cell(
                                row=cell.row, 
                                column=cell.column,
                                value=cell.value
                            )
                            try:
                                copy_cell_format(cell, new_cell)
                            except Exception as e:
                                print(f"复制单元格格式时出错: {str(e)}")
                    
                    # 复制合并单元格
                    for merged_range in source_sheet.merged_cells.ranges:
                        target_sheet.merge_cells(str(merged_range))
            
            wb.close()
        
        # 创建合并后的汇总sheet
         # 合并汇总表
        merged_summary_data = merge_summary_sheets(summary_sheets)
        
        # 写入合并后的数据
        if merged_summary_data:
            write_merged_summary(merged_wb, merged_summary_data)
        
        # 保存合并后的文件
        output_file = output_dir / 'merged_excel.xlsx'
        #
        if os.path.exists(output_file):
            os.remove(output_file)
        merged_wb.save(output_file)
        print(f"Excel文件已成功合并：{output_file}")
        
    except Exception as e:
        print(f"合并Excel文件时发生错误：{str(e)}")
    finally:
        # 清理临时文件
        for temp_file in temp_files:
            try:
                os.remove(temp_file)
            except:
                pass
        try:
            merged_wb.close()
        except:
            pass

if __name__ == '__main__':
    merge_excel_files()