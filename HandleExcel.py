import openpyxl
from pathlib import Path
from copy import copy
import os

def copy_cell_format(source_cell, target_cell):
    """复制单元格格式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def is_valid_excel(file_path):
    """验证Excel文件是否有效"""
    try:
        # 尝试打开文件
        wb = openpyxl.load_workbook(file_path, read_only=True)
        wb.close()
        return True
    except Exception as e:
        print(f"无效的Excel文件 {file_path}: {str(e)}")
        return False

def merge_excel_files():
    source_dir = Path('source/performance')
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # 创建新的工作簿
    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)  # 删除默认创建的sheet
    
    # 用于存储所有"汇总"sheet的数据和格式
    summary_cells = []
    max_summary_row = 0
    max_summary_col = 0
    
    try:
        # 获取所有Excel文件
        excel_files = list(source_dir.glob('*.xlsx')) 
        if not excel_files:
            raise Exception("source目录中没有找到Excel文件")

        # 遍历source目录下的所有excel文件
        for excel_file in excel_files:
            print(f"正在处理文件: {excel_file.name}")
            
            # 验证文件是否有效
            if not is_valid_excel(excel_file):
                print(f"跳过无效文件: {excel_file.name}")
                continue
                
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
            except Exception as e:
                print(f"无法打开文件 {excel_file.name}: {str(e)}")
                continue
            
            # 处理"汇总"sheet
            if '汇总' in wb.sheetnames:
                print(f"正在处理 {excel_file.name} 的汇总sheet")
                summary_sheet = wb['汇总']
                # 获取该sheet的使用范围
                for row in summary_sheet.rows:
                    row_data = []
                    for cell in row:
                        row_data.append({
                            'value': cell.value,
                            'cell': cell
                        })
                    summary_cells.append(row_data)
                    max_summary_col = max(max_summary_col, len(row_data))
                max_summary_row = max(max_summary_row, len(summary_cells))
            
            # 复制其他sheet
            for sheet_name in wb.sheetnames:
                if sheet_name != '汇总':
                    print(f"正在复制 {excel_file.name} 的 {sheet_name} sheet")
                    source_sheet = wb[sheet_name]
                    # 创建新的sheet，添加序号避免重名
                    new_sheet_name = f"{sheet_name}_{len(merged_wb.sheetnames)}"
                    target_sheet = merged_wb.create_sheet(title=new_sheet_name)
                    
                    # 复制列宽
                    for col in range(1, source_sheet.max_column + 1):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        if col_letter in source_sheet.column_dimensions:
                            target_sheet.column_dimensions[col_letter].width = \
                                source_sheet.column_dimensions[col_letter].width
                    
                    # 复制行高
                    for row in range(1, source_sheet.max_row + 1):
                        if row in source_sheet.row_dimensions:
                            target_sheet.row_dimensions[row].height = \
                                source_sheet.row_dimensions[row].height
                    
                    # 复制单元格内容和格式
                    for row in source_sheet.rows:
                        for cell in row:
                            new_cell = target_sheet.cell(
                                row=cell.row, 
                                column=cell.column,
                                value=cell.value
                            )
                            try:
                                copy_cell_format(cell, new_cell)
                            except Exception as e:
                                print(f"复制单元格格式时出错: {str(e)}")
                    
                    # 复制合并单元格
                    for merged_range in source_sheet.merged_cells.ranges:
                        target_sheet.merge_cells(str(merged_range))
            
            wb.close()
        
        # 创建合并后的汇总sheet
        if summary_cells:
            print("正在创建合并后的汇总sheet")
            summary_sheet = merged_wb.create_sheet(title='汇总', index=0)
            
            # 写入汇总数据和格式
            for row_idx, row_data in enumerate(summary_cells, 1):
                for col_idx, cell_data in enumerate(row_data, 1):
                    new_cell = summary_sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=cell_data['value']
                    )
                    try:
                        copy_cell_format(cell_data['cell'], new_cell)
                    except Exception as e:
                        print(f"复制汇总sheet格式时出错: {str(e)}")
        
        # 保存合并后的文件
        output_file = output_dir / 'merged_excel.xlsx'
        merged_wb.save(output_file)
        print(f"Excel文件已成功合并：{output_file}")
        
    except Exception as e:
        print(f"合并Excel文件时发生错误：{str(e)}")
    finally:
        try:
            merged_wb.close()
        except:
            pass

if __name__ == '__main__':
    merge_excel_files()